import streamlit as st
import pandas as pd
import os
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ------------------------------
# CONFIGURATION GLOBALE
# ------------------------------
st.set_page_config(page_title="Analyse Planning EHPAD Cante Cigale", layout="wide")

EXPORT_DIR = "Exports"
os.makedirs(EXPORT_DIR, exist_ok=True)

# Codes horaires et durées
CODES_HORAIRES = {
    "JWE": 10.25,
    "MA": 7.5, "M1": 7.5, "M2": 7.5, "M3": 7.5, "M4": 7.5, "S": 7.5, "SA": 7.5,
    "SE": 7.25,
    "N": 10,
    "815": 10
}

# ------------------------------
# FONCTIONS UTILES
# ------------------------------
def calculer_heures(df):
    resultats = []
    for _, row in df.iterrows():
        nom = row.iloc[0]
        for jour, code in row.iloc[1:].items():
            if isinstance(code, str) and code.strip():
                heures = CODES_HORAIRES.get(code.strip(), 0)
                resultats.append({"Nom": nom, "Jour": jour, "Code": code, "Heures": heures})
    return pd.DataFrame(resultats)

def appliquer_code_couleur(val, moyenne):
    if val < moyenne:
        return (220, 220, 220)  # Gris clair
    elif abs(val - moyenne) / moyenne <= 0.1:
        return (200, 255, 200)  # Vert clair
    elif (val - moyenne) / moyenne <= 0.25:
        return (255, 230, 200)  # Orange clair
    else:
        return (255, 180, 180)  # Rouge clair

def generer_pdf(titre, tableau, fichier_pdf):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, "Analyse Planning – EHPAD Cante Cigale", ln=True, align="C")
    pdf.set_font("Arial", 'B', 14)
    pdf.cell(200, 10, titre, ln=True, align="C")
    pdf.ln(10)

    moyenne_globale = tableau["Moyenne mensuelle"].mean()

    pdf.set_font("Arial", size=9)
    col_width = pdf.w / (len(tableau.columns) + 1)

    # En-tête
    for col in tableau.columns:
        pdf.set_fill_color(180, 180, 180)
        pdf.cell(col_width, 8, str(col), border=1, align="C", fill=True)
    pdf.ln()

    # Lignes avec couleur
    for _, row in tableau.iterrows():
        for col in tableau.columns:
            val = row[col]
            if isinstance(val, (int, float)):
                bg = appliquer_code_couleur(val, moyenne_globale)
                pdf.set_fill_color(*bg)
                txt = f"{val:.1f}"
            else:
                pdf.set_fill_color(255, 255, 255)
                txt = str(val)
            pdf.cell(col_width, 8, txt, border=1, align="C", fill=True)
        pdf.ln()

    pdf.ln(5)
    pdf.set_font("Arial", 'I', 9)
    pdf.multi_cell(0, 6,
        "Légende des couleurs :\n"
        "🟩 Vert : dans la moyenne ±10%\n"
        "🟧 Orange : +10% à +25% au-dessus de la moyenne\n"
        "🟥 Rouge : >25% au-dessus de la moyenne\n"
        "⬜ Gris : en dessous de la moyenne"
    )

    pdf.output(fichier_pdf)

def appliquer_couleurs_excel(fichier_excel, moyenne):
    wb = load_workbook(fichier_excel)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                r, g, b = appliquer_code_couleur(cell.value, moyenne)
                couleur_hex = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=couleur_hex, end_color=couleur_hex, fill_type="solid")
    wb.save(fichier_excel)

# ------------------------------
# INTERFACE STREAMLIT
# ------------------------------
st.sidebar.title("Menu")
choix = st.sidebar.radio("Choisir une analyse :", ["Analyse mensuelle", "Tableau de bord multi-mois"])

# ------------------------------
# ANALYSE MENSUELLE
# ------------------------------
if choix == "Analyse mensuelle":
    st.title("Analyse mensuelle du planning")

    fichier = st.file_uploader("Téléverser le fichier Excel du planning", type=["xlsx"])
    if fichier:
        df = pd.read_excel(fichier)
        resultats = calculer_heures(df)

        if not resultats.empty:
            total_par_nom = resultats.groupby("Nom")["Heures"].sum().reset_index()
            st.subheader("Totaux par agent")
            st.dataframe(total_par_nom)

            # Graphique
            fig, ax = plt.subplots()
            sns.barplot(x="Heures", y="Nom", data=total_par_nom.sort_values("Heures", ascending=False), ax=ax)
            ax.set_title("Répartition des heures - Analyse mensuelle")
            st.pyplot(fig)

            # Export Excel
            nom_excel = os.path.join(EXPORT_DIR, f"Analyse_Mensuelle_{datetime.date.today()}.xlsx")
            with pd.ExcelWriter(nom_excel, engine="openpyxl") as writer:
                total_par_nom.to_excel(writer, index=False, sheet_name="Synthese")
            st.success(f"Fichier Excel généré : {nom_excel}")

            # Export PDF
            nom_pdf = os.path.join(EXPORT_DIR, f"Analyse_Mensuelle_{datetime.date.today()}.pdf")
            generer_pdf("Synthèse Analyse Mensuelle", total_par_nom, nom_pdf)
            st.success(f"Fichier PDF généré : {nom_pdf}")

# ------------------------------
# TABLEAU DE BORD MULTI-MOIS
# ------------------------------
if choix == "Tableau de bord multi-mois":
    st.title("Tableau de bord multi-mois")

    fichiers = st.file_uploader("Téléverser plusieurs fichiers Excel (jusqu'à 12)", type=["xlsx"], accept_multiple_files=True)

    if fichiers:
        synthese_globale = pd.DataFrame()

        for fichier in fichiers:
            mois_annee = os.path.splitext(fichier.name)[0]
            df = pd.read_excel(fichier)
            resultats = calculer_heures(df)
            total = resultats.groupby("Nom")["Heures"].sum().reset_index()
            total.rename(columns={"Heures": mois_annee}, inplace=True)

            if synthese_globale.empty:
                synthese_globale = total
            else:
                synthese_globale = pd.merge(synthese_globale, total, on="Nom", how="outer")

        synthese_globale.fillna(0, inplace=True)

        synthese_globale["Total annuel"] = synthese_globale.iloc[:, 1:].sum(axis=1)
        synthese_globale["Moyenne mensuelle"] = synthese_globale.iloc[:, 1:-1].mean(axis=1)

        st.subheader("Synthèse multi-mois")
        st.dataframe(synthese_globale)

        # Graphique évolution
        st.subheader("Évolution par agent")
        fig, ax = plt.subplots(figsize=(10,6))
        for _, row in synthese_globale.iterrows():
            ax.plot(synthese_globale.columns[1:-2], row[1:-2], marker="o", label=row["Nom"])
        ax.set_xlabel("Mois")
        ax.set_ylabel("Heures")
        ax.set_title("Évolution des heures dimanches/fériés")
        ax.legend()
        st.pyplot(fig)

        # Export Excel avec couleurs
        nom_excel = os.path.join(EXPORT_DIR, f"Comparatif_Annuel_{datetime.date.today()}.xlsx")
        with pd.ExcelWriter(nom_excel, engine="openpyxl") as writer:
            synthese_globale.to_excel(writer, index=False, sheet_name="Comparatif Annuel")
        appliquer_couleurs_excel(nom_excel, synthese_globale["Moyenne mensuelle"].mean())
        st.success(f"Fichier Excel généré : {nom_excel}")

        # Export PDF avec couleurs
        nom_pdf = os.path.join(EXPORT_DIR, f"Comparatif_Annuel_{datetime.date.today()}.pdf")
        generer_pdf("Synthèse Comparatif Multi-mois", synthese_globale, nom_pdf)
        st.success(f"Fichier PDF généré : {nom_pdf}")
